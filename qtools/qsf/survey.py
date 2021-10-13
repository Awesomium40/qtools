from openpyxl.cell.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from .abc import SurveyObjectBase
from .questions import *
import openpyxl


class _Survey(SurveyObjectBase):

    COL_HEADERS = ["QuestionID", 'Variable Name', "Variable Label", "Value Labels"]
    WIDTHS = {"A": 12.0, "B": 16.0, "C": 60.0, "D": 40.0}
    ALIGNMENTS = {"A": openpyxl.styles.Alignment(),
                  "B": openpyxl.styles.Alignment(),
                  "C": openpyxl.styles.Alignment(wrapText=True),
                  "D": openpyxl.styles.Alignment(wrapText=True)}

    TBL_STYLE = TableStyleInfo(name="TableStyleMedium8",
                               showRowStripes=True)

    _question_map_ = {'Matrix': MatrixQuestion,
                      'MC': MultiChoiceQuestion,
                      "RO": RankOrderQuestion,
                      "SQ": SliderQuestion,
                      "TE": TextEntryQuestion,
                      'SBS': SideBySideQuestion}

    def __init__(self, items, **kwargs):
        super().__init__(items, **kwargs)

    def get_question(self, question_id):
        """
        s.get_question(question_id) -> SurveyQuestion
        finds the question with ID question_id and returns it
        :param question_id: the ID of the question to find
        :return: SurveyQuestion
        :raises ValueError: if no question can be found
        """
        question = next(filter(lambda x: x.Element == 'SQ' and x.Payload.QuestionID == question_id,
                               self.SurveyElements), None)

        if question is None:
            raise ValueError(f'Question {question_id} not found')

        return question

    def __xl__codebook(self):

        blocks = next(filter(lambda x: x.get('Element') == 'BL', self.SurveyElements))
        payload = blocks.Payload
        block_items = filter(lambda x: x.get('Type') in ("Default", "Standard"),
                             payload if isinstance(payload, list) else payload.values())

        trash = next(filter(lambda x: x.get('Type') == 'Trash', payload if isinstance(payload, list)
                            else payload.values()))
        trash_ids = [x.QuestionID for x in trash.BlockElements if hasattr(x, 'QuestionID')]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self.COL_HEADERS)

        for i, block in enumerate(block_items):
            title = block.Description

            for key, value in self.WIDTHS.items():
                ws.column_dimensions[key].width = value

            # Add the title of each block and merge the cells
            ws.append([title])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)

            questions = filter(lambda x: x.Type == "Question" and x.QuestionID not in trash_ids,
                               block.BlockElements)

            for q in questions:

                question_id = q.QuestionID
                question = next(filter(lambda x: x.Element == 'SQ' and x.Payload.QuestionID == question_id,
                                       self.SurveyElements))
                try:
                    for row in question.variable_info():

                        # If there are value labels, they need to go into their own row and have the rest of the var
                        # info in merged cells occupying the same number of rows pre-merge
                        if row[3] is not None:
                            start_row = ws.max_row + 1
                            for i, (key, value) in enumerate(sorted(row[3].items())):
                                ws.append([x.strip() for x in row[:3]] + [f'{key} = {value}'])

                            # Perform the merging on each individual cell so they are lined up with the val labels
                            for c in range(1, 4):
                                ws.merge_cells(start_row=start_row, end_row=start_row + i, start_column=c, end_column=c)
                        else:
                            ws.append(row)
                except Exception as err:
                    print(f"Unable to insert data for question {question_id}. Error was \n{err}")

            # Make a table
            """
            max_row = ws.max_row
            if max_row > 2:
                tbl = Table(displayName=title, ref=f"A1:D{max_row}")
                tbl.tableStyleInfo = self.TBL_STYLE
                ws.add_table(tbl)
            """

            # Styles have to be set on cells one at a time or they won't take
            for row_index, row in enumerate(ws.iter_rows(), 1):
                for column_index, cell in enumerate(row, 1):
                    column_letter = get_column_letter(column_index)
                    ws.cell(row_index, column_index).alignment = self.ALIGNMENTS.get(column_letter,
                                                                                     openpyxl.styles.Alignment())

        return wb

    def codebook(self):
        """

        :param output_type:
        :param var_info_only:
        :return:
        """
        return self.__xl__codebook()

