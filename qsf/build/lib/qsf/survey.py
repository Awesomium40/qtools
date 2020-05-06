from openpyxl.cell.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from .abc import SurveyObject
import openpyxl


class _Survey(SurveyObject):
    COL_HEADERS = ['QuestionID', 'QuestionType', 'Selector', 'SubSelector', 'QuestionStem', 'QuestionLeaf',
                   'Variable Name', 'Variable Label', 'Value Labels', 'Validation']
    WIDTHS = {"A": 10.3, "B": 21.5, "C": 44.3, "D": 27.9}
    ALIGNMENTS = {"A": openpyxl.styles.Alignment(),
                  "B": openpyxl.styles.Alignment(),
                  "C": openpyxl.styles.Alignment(wrapText=True),
                  "D": openpyxl.styles.Alignment(wrapText=True)}

    TBL_STYLE = TableStyleInfo(name="TableStyleMedium8",
                               showRowStripes=True)

    def __init__(self, items, **kwargs):

        super().__init__(items, **kwargs)

    def __docx_codebook(self, var_info_only=True):
        raise NotImplementedError()

    def __xl__codebook(self, var_info_only=True):

        col_headers = _Survey.COL_HEADERS if not var_info_only else [_Survey.COL_HEADERS[0]] + _Survey.COL_HEADERS[6:9]
        flow_elements = next(filter(lambda x: x.get('Element') == 'FL', self.SurveyElements)).Payload.Flow
        blocks = next(filter(lambda x: x.get('Element') == 'BL', self.SurveyElements))
        payload = blocks.Payload
        block_items = filter(lambda x: x.get('Type') in ("Default", "Standard"),
                             payload if isinstance(payload, list) else payload.values())

        trash = next(filter(lambda x: x.get('Type') == 'Trash', payload if isinstance(payload, list)
                            else payload.values()))
        trash_ids = [x.QuestionID for x in trash.BlockElements if hasattr(x, 'QuestionID')]
        wb = openpyxl.Workbook()

        # TODO: Instead of enumerating over flow elements, iterate over block elements
        for i, block in enumerate(block_items):
            """
        for i, flow in enumerate(flow_elements if isinstance(flow_elements, list) else flow_elements.values()):
            block = next(filter(lambda x: x.ID == flow.ID, blocks.Payload), None)
            """
            ws = wb.active if i == 0 else wb.create_sheet()
            title = self._sanitize_for_spss_(block.Description)
            title = title if len(title) < 32 else title[:31]
            ws.title = title
            for key, value in self.WIDTHS.items():
                ws.column_dimensions[key].width = value

            ws.append(col_headers)
            questions = filter(lambda x: x.Type == "Question" and x.QuestionID not in trash_ids,
                               block.BlockElements)

            for q in questions:
                question_id = q.QuestionID
                if question_id == 'QID21':
                    x = "Toasty"
                question = next(filter(lambda x: x.Element == 'SQ' and x.Payload.QuestionID == question_id,
                                       self.SurveyElements))
                try:
                    for row in question.variable_info():
                        row_data = row if not var_info_only else [row[0]] + list(row[6:9])
                        ws.append(row_data)
                except Exception as err:
                    print("Unable to insert data for question {0}. Error was \n{1}".format(question_id, err))

            # Make a table
            max_row = ws.max_row
            if max_row > 1:
                tbl = Table(displayName=title, ref="A1:D{0}".format(max_row))
                tbl.tableStyleInfo = self.TBL_STYLE
                ws.add_table(tbl)

            # Styles have to be set on cells one at a time or they won't take
            for row_index, row in enumerate(ws.iter_rows(), 1):
                for column_index, cell in enumerate(row, 1):
                    column_letter = get_column_letter(column_index)
                    ws.cell(row_index, column_index).alignment = self.ALIGNMENTS.get(column_letter,
                                                                                     openpyxl.styles.Alignment())

        return wb

    def codebook(self, output_type='xlsx', var_info_only=True):
        """

        :param output_type:
        :param var_info_only:
        :return:
        """

        if output_type == 'xlsx':
            return self.__xl__codebook(var_info_only=var_info_only)
        else:
            raise NotImplementedError("The output type '{0}' is not currently supported".format(output_type))
